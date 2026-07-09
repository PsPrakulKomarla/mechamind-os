"""
Document Ingestion Pipeline
Processes various document types and extracts text, chunks, embeddings, and entities.
"""

import os
import asyncio
from typing import List, Dict, Any, Optional
from uuid import UUID
from datetime import datetime
from pathlib import Path
import json

import pdfplumber
import pytesseract
from PIL import Image
from pdf2image import convert_from_path
import openpyxl
from docx import Document as DocxDocument
from unstructured.partition.auto import partition
from unstructured.chunking.title import chunk_by_title

from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select

from app.db.models import Document, DocumentChunk, DocumentType, DocumentStatus, Entity, EntityRelationship
from app.services.rag.embeddings import EmbeddingService
from app.services.vector_store.qdrant_client import QdrantVectorStore
from app.services.knowledge_graph.extractor import EntityExtractor
from app.config import settings


class IngestionPipeline:
    """Main pipeline for processing uploaded documents."""
    
    def __init__(self):
        self.embedding_service = EmbeddingService()
        self.vector_store = QdrantVectorStore()
        self.entity_extractor = EntityExtractor()
    
    async def process_document(self, document_id: UUID, file_path: str, doc_type: DocumentType):
        """Process a document through the full pipeline."""
        from app.db.database import async_session_maker
        
        async with async_session_maker() as db:
            # Get document
            result = await db.execute(select(Document).where(Document.id == document_id))
            document = result.scalar_one_or_none()
            
            if not document:
                raise ValueError(f"Document {document_id} not found")
            
            try:
                document.status = DocumentStatus.PROCESSING
                await db.commit()
                
                # Step 1: Extract text based on document type
                print(f"📄 Extracting text from {document_id}...")
                extracted = await self._extract_text(file_path, doc_type)
                
                # Step 2: Chunk the text
                print(f"✂️ Chunking text...")
                chunks = await self._chunk_text(extracted["text"], extracted.get("metadata", {}))
                
                # Step 3: Generate embeddings
                print(f"🔢 Generating embeddings for {len(chunks)} chunks...")
                embedded_chunks = await self._embed_chunks(chunks)
                
                # Step 4: Store chunks in database and vector store
                print(f"💾 Storing chunks...")
                await self._store_chunks(db, document_id, embedded_chunks)
                
                # Step 5: Extract entities and relationships
                print(f"🔍 Extracting entities...")
                entities = await self.entity_extractor.extract_from_document(
                    document_id, extracted["text"]
                )
                await self._store_entities(db, document_id, entities)
                
                # Step 6: Update document status
                document.status = DocumentStatus.READY
                document.processed_at = datetime.utcnow()
                document.page_count = extracted.get("page_count")
                document.metadata = {
                    **(document.metadata or {}),
                    "extracted_pages": extracted.get("page_count"),
                    "total_chunks": len(chunks),
                    "total_entities": len(entities),
                    "processing_metadata": extracted.get("metadata", {}),
                }
                
                await db.commit()
                print(f"✅ Document {document_id} processed successfully")
                
            except Exception as e:
                document.status = DocumentStatus.FAILED
                document.error_message = str(e)
                await db.commit()
                print(f"❌ Document {document_id} processing failed: {e}")
                raise
    
    async def _extract_text(self, file_path: str, doc_type: DocumentType) -> Dict[str, Any]:
        """Extract text from various document types."""
        
        if doc_type == DocumentType.PDF:
            return await self._extract_pdf(file_path)
        elif doc_type == DocumentType.IMAGE:
            return await self._extract_image(file_path)
        elif doc_type == DocumentType.SPREADSHEET:
            return await self._extract_spreadsheet(file_path)
        elif doc_type == DocumentType.TEXT:
            return await self._extract_text_file(file_path)
        elif doc_type == DocumentType.EMAIL:
            return await self._extract_email(file_path)
        else:
            # Fallback to unstructured
            return await self._extract_generic(file_path)
    
    async def _extract_pdf(self, file_path: str) -> Dict[str, Any]:
        """Extract text from PDF with OCR fallback."""
        text_parts = []
        metadata = {"pages": []}
        
        try:
            # Try pdfplumber first (fast, good for text PDFs)
            with pdfplumber.open(file_path) as pdf:
                metadata["page_count"] = len(pdf.pages)
                
                for i, page in enumerate(pdf.pages):
                    page_text = page.extract_text()
                    if page_text and page_text.strip():
                        text_parts.append(page_text)
                        metadata["pages"].append({
                            "page": i + 1,
                            "chars": len(page_text),
                            "method": "pdfplumber"
                        })
                    else:
                        # Page might be scanned - needs OCR
                        metadata["pages"].append({
                            "page": i + 1,
                            "chars": 0,
                            "method": "needs_ocr"
                        })
            
            # If no text extracted, or some pages need OCR
            if not text_parts or any(p["method"] == "needs_ocr" for p in metadata["pages"]):
                print("  📸 PDF appears scanned, running OCR...")
                ocr_result = await self._ocr_pdf(file_path)
                text_parts = ocr_result["text_parts"]
                metadata = ocr_result["metadata"]
        
        except Exception as e:
            print(f"  ⚠️ pdfplumber failed: {e}, trying OCR...")
            ocr_result = await self._ocr_pdf(file_path)
            text_parts = ocr_result["text_parts"]
            metadata = ocr_result["metadata"]
        
        return {
            "text": "\n\n".join(text_parts),
            "metadata": metadata,
            "page_count": metadata.get("page_count"),
        }
    
    async def _ocr_pdf(self, file_path: str) -> Dict[str, Any]:
        """Run OCR on PDF pages."""
        text_parts = []
        metadata = {"pages": []}
        
        # Convert PDF to images
        images = convert_from_path(file_path, dpi=200)
        metadata["page_count"] = len(images)
        
        for i, image in enumerate(images):
            # Preprocess image for better OCR
            processed = self._preprocess_for_ocr(image)
            
            # Run OCR
            page_text = pytesseract.image_to_string(processed, lang='eng')
            
            if page_text.strip():
                text_parts.append(page_text)
                metadata["pages"].append({
                    "page": i + 1,
                    "chars": len(page_text),
                    "method": "tesseract_ocr"
                })
            else:
                metadata["pages"].append({
                    "page": i + 1,
                    "chars": 0,
                    "method": "tesseract_ocr_empty"
                })
        
        return {"text_parts": text_parts, "metadata": metadata}
    
    def _preprocess_for_ocr(self, image: Image.Image) -> Image.Image:
        """Preprocess image for better OCR results."""
        # Convert to grayscale
        if image.mode != 'L':
            image = image.convert('L')
        
        # Increase contrast
        from PIL import ImageEnhance
        enhancer = ImageEnhance.Contrast(image)
        image = enhancer.enhance(2.0)
        
        # Sharpen
        enhancer = ImageEnhance.Sharpness(image)
        image = enhancer.enhance(1.5)
        
        return image
    
    async def _extract_image(self, file_path: str) -> Dict[str, Any]:
        """Extract text from image using OCR."""
        image = Image.open(file_path)
        processed = self._preprocess_for_ocr(image)
        text = pytesseract.image_to_string(processed, lang='eng')
        
        return {
            "text": text,
            "metadata": {"page_count": 1, "pages": [{"page": 1, "chars": len(text), "method": "tesseract_ocr"}]},
            "page_count": 1,
        }
    
    async def _extract_spreadsheet(self, file_path: str) -> Dict[str, Any]:
        """Extract text from spreadsheet."""
        wb = openpyxl.load_workbook(file_path, data_only=True)
        text_parts = []
        metadata = {"sheets": []}
        
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            sheet_text = []
            
            for row in ws.iter_rows(values_only=True):
                row_text = " | ".join(str(c) if c is not None else "" for c in row)
                if row_text.strip():
                    sheet_text.append(row_text)
            
            if sheet_text:
                text_parts.append(f"Sheet: {sheet_name}\n" + "\n".join(sheet_text))
                metadata["sheets"].append({
                    "name": sheet_name,
                    "rows": ws.max_row,
                    "cols": ws.max_column,
                })
        
        return {
            "text": "\n\n".join(text_parts),
            "metadata": metadata,
            "page_count": len(wb.sheetnames),
        }
    
    async def _extract_text_file(self, file_path: str) -> Dict[str, Any]:
        """Extract text from plain text file."""
        with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
            text = f.read()
        
        return {
            "text": text,
            "metadata": {"page_count": 1},
            "page_count": 1,
        }
    
    async def _extract_email(self, file_path: str) -> Dict[str, Any]:
        """Extract text from email file."""
        # Use unstructured for email parsing
        elements = partition(filename=file_path)
        text = "\n\n".join([str(el) for el in elements])
        
        return {
            "text": text,
            "metadata": {"page_count": 1},
            "page_count": 1,
        }
    
    async def _extract_generic(self, file_path: str) -> Dict[str, Any]:
        """Generic extraction using unstructured."""
        elements = partition(filename=file_path)
        text = "\n\n".join([str(el) for el in elements])
        
        return {
            "text": text,
            "metadata": {"page_count": 1},
            "page_count": 1,
        }
    
    async def _chunk_text(self, text: str, metadata: Dict) -> List[Dict[str, Any]]:
        """Split text into chunks with metadata."""
        # Use unstructured's chunking for semantic chunks
        elements = partition(text=text)
        chunks = chunk_by_title(elements, max_characters=settings.CHUNK_SIZE, overlap=settings.CHUNK_OVERLAP)
        
        chunk_data = []
        for i, chunk in enumerate(chunks):
            chunk_text = str(chunk)
            if len(chunk_text.strip()) < 50:  # Skip very small chunks
                continue
            
            chunk_meta = {
                "chunk_index": i,
                "source_metadata": metadata,
                "element_type": type(chunk).__name__,
            }
            
            # Extract page number if available
            if hasattr(chunk, 'metadata') and chunk.metadata.page_number:
                chunk_meta["page_number"] = chunk.metadata.page_number
            
            chunk_data.append({
                "content": chunk_text,
                "metadata": chunk_meta,
                "token_count": len(chunk_text) // 4,  # Rough estimate
            })
        
        return chunk_data
    
    async def _embed_chunks(self, chunks: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
        """Generate embeddings for chunks."""
        texts = [c["content"] for c in chunks]
        embeddings = await self.embedding_service.embed_batch(texts)
        
        for chunk, embedding in zip(chunks, embeddings):
            chunk["embedding"] = embedding
        
        return chunks
    
    async def _store_chunks(self, db: AsyncSession, document_id: UUID, chunks: List[Dict[str, Any]]):
        """Store chunks in PostgreSQL and vector store."""
        # Store in PostgreSQL
        for chunk in chunks:
            db_chunk = DocumentChunk(
                document_id=document_id,
                chunk_index=chunk["metadata"]["chunk_index"],
                content=chunk["content"],
                embedding=chunk["embedding"],
                metadata=chunk["metadata"],
                token_count=chunk.get("token_count"),
            )
            db.add(db_chunk)
        
        await db.flush()
        
        # Store in vector store
        vector_points = []
        for chunk in chunks:
            # Get the DB chunk ID (need to query or track)
            # For now, we'll upsert by content hash
            vector_points.append({
                "id": str(UUID(bytes=hash(chunk["content"].encode())[:16])),
                "vector": chunk["embedding"],
                "payload": {
                    "document_id": str(document_id),
                    "content": chunk["content"][:1000],  # Truncate for payload
                    "metadata": chunk["metadata"],
                }
            })
        
        await self.vector_store.upsert_batch(vector_points)
    
    async def _store_entities(self, db: AsyncSession, document_id: UUID, entities: List[Dict]):
        """Store extracted entities and relationships."""
        entity_map = {}
        
        for ent in entities:
            db_entity = Entity(
                name=ent["name"],
                type=ent["type"],
                properties=ent.get("properties", {}),
                description=ent.get("description"),
                source_document_id=document_id,
                confidence=ent.get("confidence", 0.8),
            )
            db.add(db_entity)
            await db.flush()
            entity_map[ent["name"]] = db_entity.id
        
        # Store relationships
        for ent in entities:
            if "relationships" in ent:
                for rel in ent["relationships"]:
                    if rel["target"] in entity_map:
                        db_rel = EntityRelationship(
                            source_id=entity_map[ent["name"]],
                            target_id=entity_map[rel["target"]],
                            relationship_type=rel["type"],
                            properties=rel.get("properties", {}),
                            confidence=rel.get("confidence", 0.7),
                            source="extracted",
                        )
                        db.add(db_rel)
        
        await db.commit()


# Global pipeline instance
ingestion_pipeline = IngestionPipeline()